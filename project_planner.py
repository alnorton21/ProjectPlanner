"""
Project Planner - A comprehensive project management application
Modeled after Microsoft Teams Planner with Gantt charts and Excel export
Structure: Project -> Activity -> Work Item (multiple employees at each tier)
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import json
import os
import uuid
from datetime import datetime, date, timedelta
from typing import Optional
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.dates as mdates
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import openpyxl.chart.series
import colorsys

# ─────────────────────────────────────────────
# DATA MODEL
# ─────────────────────────────────────────────

STATUS_OPTIONS = ["Not Started", "In Progress", "Completed", "Blocked", "On Hold"]
PRIORITY_OPTIONS = ["Low", "Medium", "High", "Critical"]
STATUS_COLORS = {
    "Not Started": "#94a3b8",
    "In Progress": "#3b82f6",
    "Completed": "#22c55e",
    "Blocked": "#ef4444",
    "On Hold": "#f59e0b",
}
PRIORITY_COLORS = {
    "Low": "#94a3b8",
    "Medium": "#f59e0b",
    "High": "#f97316",
    "Critical": "#ef4444",
}

PALETTE = [
    "#3b82f6", "#8b5cf6", "#ec4899", "#14b8a6",
    "#f97316", "#22c55e", "#eab308", "#06b6d4",
    "#6366f1", "#10b981", "#f43f5e", "#84cc16",
]


def today_str():
    return date.today().isoformat()


def new_id():
    return str(uuid.uuid4())[:8]


def _to_list(v):
    """Coerce a value to a list — handles migration from old string format."""
    if isinstance(v, list):
        return v
    if v:
        return [v]
    return []


class WorkItem:
    def __init__(self, title="", description="", assigned_to=None,
                 status="Not Started", priority="Medium",
                 start_date=None, due_date=None, progress=0,
                 checklist=None, item_id=None):
        self.id = item_id or new_id()
        self.title = title
        self.description = description
        self.assigned_to = _to_list(assigned_to)   # list of employee names
        self.status = status
        self.priority = priority
        self.start_date = start_date or today_str()
        self.due_date = due_date or today_str()
        self.progress = progress
        self.checklist = checklist or []

    def to_dict(self):
        return self.__dict__.copy()

    @classmethod
    def from_dict(cls, d):
        d = d.copy()
        item_id = d.pop("id", None)
        return cls(item_id=item_id, **d)


class Activity:
    def __init__(self, title="", description="", employees=None,
                 priority="Medium", start_date=None, due_date=None, act_id=None):
        self.id = act_id or new_id()
        self.title = title
        self.description = description
        self.employees = _to_list(employees)
        self.priority = priority
        self.start_date = start_date or today_str()
        self.due_date = due_date or today_str()
        self.work_items: list[WorkItem] = []

    @property
    def status(self):
        """Computed from work items — cannot be set manually."""
        if not self.work_items:
            return "Not Started"
        statuses = [wi.status for wi in self.work_items]
        s = set(statuses)
        if s == {"Completed"}:
            return "Completed"
        if "Blocked" in s:
            return "Blocked"
        if s == {"Not Started"}:
            return "Not Started"
        if s <= {"On Hold", "Not Started"}:
            return "On Hold"
        return "In Progress"

    @property
    def progress(self):
        """Computed as average of work item progress values."""
        if not self.work_items:
            return 0
        return int(sum(w.progress for w in self.work_items) / len(self.work_items))

    def to_dict(self):
        return {
            "id": self.id,
            "title": self.title,
            "description": self.description,
            "employees": self.employees,
            "priority": self.priority,
            "start_date": self.start_date,
            "due_date": self.due_date,
            "work_items": [w.to_dict() for w in self.work_items],
        }

    @classmethod
    def from_dict(cls, d):
        d = d.copy()
        items = d.pop("work_items", [])
        act_id = d.pop("id", None)
        # drop computed/legacy fields
        d.pop("status", None)
        d.pop("progress", None)
        if "assigned_to" in d and "employees" not in d:
            d["employees"] = _to_list(d.pop("assigned_to"))
        elif "assigned_to" in d:
            d.pop("assigned_to")
        act = cls(act_id=act_id, **{k: v for k, v in d.items()})
        act.work_items = [WorkItem.from_dict(w) for w in items]
        return act


class Project:
    def __init__(self, title="New Project", description="",
                 start_date=None, due_date=None, proj_id=None):
        self.id = proj_id or new_id()
        self.title = title
        self.description = description
        self.start_date = start_date or today_str()
        self.due_date = due_date or today_str()
        self.activities: list[Activity] = []
        self.employees: list[dict] = []   # {"name", "role", "email"}
        self.color = PALETTE[0]

    @property
    def progress(self):
        if not self.activities:
            return 0
        return int(sum(a.progress for a in self.activities) / len(self.activities))

    def to_dict(self):
        d = {k: v for k, v in self.__dict__.items() if k != "activities"}
        d["activities"] = [a.to_dict() for a in self.activities]
        return d

    @classmethod
    def from_dict(cls, d):
        d = d.copy()
        # support both new (activities) and old (subtasks) save format
        acts_raw = d.pop("activities", None)
        if acts_raw is None:
            old_subs = d.pop("subtasks", [])
            acts_raw = []
            for sub in old_subs:
                sub = sub.copy()
                objs = sub.pop("objectives", [])
                work_items = []
                for obj in objs:
                    work_items.extend(obj.get("work_items", []))
                sub["work_items"] = work_items
                acts_raw.append(sub)
        employees = d.pop("employees", [])
        color = d.pop("color", PALETTE[0])
        skip = {"id", "activities", "employees", "color"}
        p = cls(proj_id=d.get("id"), **{k: v for k, v in d.items() if k not in skip})
        p.activities = [Activity.from_dict(a) for a in acts_raw]
        p.employees = employees
        p.color = color
        return p


class AppData:
    def __init__(self):
        self.projects: list[Project] = []
        self.global_employees: list[dict] = []

    def save(self, path):
        data = {
            "projects": [p.to_dict() for p in self.projects],
            "global_employees": self.global_employees,
        }
        with open(path, "w") as f:
            json.dump(data, f, indent=2)

    def load(self, path):
        with open(path) as f:
            data = json.load(f)
        self.projects = [Project.from_dict(p) for p in data.get("projects", [])]
        self.global_employees = data.get("global_employees", [])

    def all_employees(self):
        names = set()
        for p in self.projects:
            for e in p.employees:
                names.add(e["name"])
        for e in self.global_employees:
            names.add(e["name"])
        return sorted(names)


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def parse_date(s):
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return date.today()


def date_to_num(d):
    return mdates.date2num(datetime.combine(d, datetime.min.time()))


def hex_to_rgb(h):
    h = h.lstrip("#")
    return tuple(int(h[i:i+2], 16) / 255.0 for i in (0, 2, 4))


def lighten(hex_color, factor=0.4):
    r, g, b = hex_to_rgb(hex_color)
    h, s, v = colorsys.rgb_to_hsv(r, g, b)
    v = min(1.0, v + factor)
    s = max(0, s - factor * 0.5)
    r2, g2, b2 = colorsys.hsv_to_rgb(h, s, v)
    return "#{:02x}{:02x}{:02x}".format(int(r2*255), int(g2*255), int(b2*255))


# ─────────────────────────────────────────────
# WIDGETS
# ─────────────────────────────────────────────

class MultiEmployeeWidget(tk.Frame):
    """Listbox + combobox for assigning multiple employees."""
    def __init__(self, parent, available, selected=None, **kw):
        super().__init__(parent, bg="#0f172a", **kw)
        self._selected = list(selected or [])
        self._available = list(available or [])
        self._build()

    def _build(self):
        lf = tk.Frame(self, bg="#0f172a")
        lf.pack(fill="x")
        self._lb = tk.Listbox(lf, bg="#1e293b", fg="#e2e8f0",
                               selectbackground="#1e3a5f", relief="flat",
                               font=("Consolas", 9), height=3,
                               exportselection=False, borderwidth=0,
                               highlightthickness=0)
        self._lb.pack(side="left", fill="x", expand=True)
        sb = ttk.Scrollbar(lf, orient="vertical", command=self._lb.yview)
        self._lb.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")

        af = tk.Frame(self, bg="#0f172a")
        af.pack(fill="x", pady=(3, 0))
        self._add_v = tk.StringVar()
        self._combo = ttk.Combobox(af, textvariable=self._add_v,
                                    values=self._available, width=20)
        self._combo.pack(side="left")
        tk.Button(af, text="+ Add", command=self._add,
                  bg="#1e293b", fg="#3b82f6", relief="flat",
                  font=("Consolas", 8), cursor="hand2", padx=6).pack(side="left", padx=4)
        tk.Button(af, text="✕ Remove", command=self._remove,
                  bg="#1e293b", fg="#ef4444", relief="flat",
                  font=("Consolas", 8), cursor="hand2", padx=6).pack(side="left")
        self._refresh()

    def _refresh(self):
        self._lb.delete(0, "end")
        for name in self._selected:
            self._lb.insert("end", f"  {name}")

    def _add(self):
        name = self._add_v.get().strip()
        if name and name not in self._selected:
            self._selected.append(name)
            self._refresh()

    def _remove(self):
        sel = self._lb.curselection()
        if sel:
            self._selected.pop(sel[0])
            self._refresh()

    def get(self):
        return list(self._selected)


# ─────────────────────────────────────────────
# DIALOGS
# ─────────────────────────────────────────────

class BaseDialog(tk.Toplevel):
    def __init__(self, parent, title, employees=None):
        super().__init__(parent)
        self.title(title)
        self.result = None
        self.employees = employees or []
        self.grab_set()
        self.resizable(True, True)
        self._setup_style()
        self._build()
        self.wait_window()

    def _setup_style(self):
        self.configure(bg="#0f172a")
        self.option_add("*Background", "#0f172a")
        self.option_add("*Foreground", "#e2e8f0")

    def _labeled_entry(self, parent, label, row, default=""):
        tk.Label(parent, text=label, bg="#0f172a", fg="#94a3b8",
                 font=("Consolas", 9)).grid(row=row, column=0, sticky="w", padx=8, pady=3)
        v = tk.StringVar(value=default)
        e = tk.Entry(parent, textvariable=v, bg="#1e293b", fg="#e2e8f0",
                     insertbackground="#e2e8f0", relief="flat",
                     font=("Consolas", 10), width=36)
        e.grid(row=row, column=1, sticky="ew", padx=8, pady=3)
        return v

    def _labeled_combo(self, parent, label, row, options, default=""):
        tk.Label(parent, text=label, bg="#0f172a", fg="#94a3b8",
                 font=("Consolas", 9)).grid(row=row, column=0, sticky="w", padx=8, pady=3)
        v = tk.StringVar(value=default)
        c = ttk.Combobox(parent, textvariable=v, values=options,
                         state="readonly", width=33)
        c.grid(row=row, column=1, sticky="ew", padx=8, pady=3)
        return v

    def _labeled_date(self, parent, label, row, default=""):
        tk.Label(parent, text=label, bg="#0f172a", fg="#94a3b8",
                 font=("Consolas", 9)).grid(row=row, column=0, sticky="w", padx=8, pady=3)
        v = tk.StringVar(value=default or today_str())
        e = tk.Entry(parent, textvariable=v, bg="#1e293b", fg="#e2e8f0",
                     insertbackground="#e2e8f0", relief="flat",
                     font=("Consolas", 10), width=36)
        e.grid(row=row, column=1, sticky="ew", padx=8, pady=3)
        tk.Label(parent, text="(YYYY-MM-DD)", bg="#0f172a", fg="#475569",
                 font=("Consolas", 8)).grid(row=row, column=2, sticky="w", padx=4)
        return v

    def _ok_cancel(self, parent):
        bf = tk.Frame(parent, bg="#0f172a")
        bf.pack(fill="x", padx=16, pady=12)
        tk.Button(bf, text="Cancel", command=self.destroy,
                  bg="#334155", fg="#94a3b8", relief="flat",
                  font=("Consolas", 10), padx=16, pady=6,
                  cursor="hand2").pack(side="right", padx=4)
        tk.Button(bf, text="Save", command=self._save,
                  bg="#3b82f6", fg="white", relief="flat",
                  font=("Consolas", 10, "bold"), padx=16, pady=6,
                  cursor="hand2").pack(side="right", padx=4)

    def _build(self):
        raise NotImplementedError

    def _save(self):
        raise NotImplementedError


class ProjectDialog(BaseDialog):
    def __init__(self, parent, project=None, employees=None):
        self.project = project
        super().__init__(parent, "Edit Project" if project else "New Project", employees)

    def _build(self):
        self.geometry("520x400")
        tk.Label(self, text="PROJECT", bg="#0f172a", fg="#3b82f6",
                 font=("Consolas", 13, "bold")).pack(pady=(16, 8))
        f = tk.Frame(self, bg="#0f172a")
        f.pack(fill="both", expand=True, padx=8)
        f.columnconfigure(1, weight=1)
        p = self.project
        self._title_v = self._labeled_entry(f, "Title", 0, p.title if p else "")
        self._desc_v = self._labeled_entry(f, "Description", 1, p.description if p else "")
        self._start_v = self._labeled_date(f, "Start Date", 2, p.start_date if p else "")
        self._due_v = self._labeled_date(f, "Due Date", 3, p.due_date if p else "")
        tk.Label(f, text="Color", bg="#0f172a", fg="#94a3b8",
                 font=("Consolas", 9)).grid(row=4, column=0, sticky="w", padx=8, pady=3)
        cf = tk.Frame(f, bg="#0f172a")
        cf.grid(row=4, column=1, sticky="w", padx=8)
        self._color_v = tk.StringVar(value=p.color if p else PALETTE[0])
        for i, c in enumerate(PALETTE):
            tk.Button(cf, bg=c, width=2, height=1, relief="flat",
                      cursor="hand2",
                      command=lambda col=c: self._color_v.set(col)).grid(row=0, column=i, padx=1)
        self._ok_cancel(self)

    def _save(self):
        t = self._title_v.get().strip()
        if not t:
            messagebox.showerror("Error", "Title is required", parent=self)
            return
        if self.project:
            self.project.title = t
            self.project.description = self._desc_v.get()
            self.project.start_date = self._start_v.get()
            self.project.due_date = self._due_v.get()
            self.project.color = self._color_v.get()
            self.result = self.project
        else:
            p = Project(title=t, description=self._desc_v.get(),
                        start_date=self._start_v.get(),
                        due_date=self._due_v.get())
            p.color = self._color_v.get()
            self.result = p
        self.destroy()


class TaskDialog(BaseDialog):
    """Dialog for Activity or WorkItem — both support multiple employee assignment."""
    def __init__(self, parent, level="activity", item=None, employees=None):
        self.level = level
        self.item = item
        self.checklist_items = []
        labels = {"activity": "ACTIVITY", "workitem": "WORK ITEM"}
        title = f"Edit {labels.get(level, level)}" if item else f"New {labels.get(level, level)}"
        super().__init__(parent, title, employees)

    def _build(self):
        self.geometry("580x600")
        labels = {"activity": "ACTIVITY", "workitem": "WORK ITEM"}
        colors = {"activity": "#8b5cf6", "workitem": "#f97316"}
        tk.Label(self, text=labels.get(self.level, "TASK"),
                 bg="#0f172a", fg=colors.get(self.level, "#3b82f6"),
                 font=("Consolas", 13, "bold")).pack(pady=(16, 8))

        canvas = tk.Canvas(self, bg="#0f172a", highlightthickness=0)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        self.scroll_frame = tk.Frame(canvas, bg="#0f172a")
        self.scroll_frame.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        f = self.scroll_frame
        f.columnconfigure(1, weight=1)
        it = self.item
        row = 0

        self._title_v = self._labeled_entry(f, "Title", row, it.title if it else ""); row += 1
        self._desc_v = self._labeled_entry(f, "Description", row, it.description if it else ""); row += 1

        emp_label = "Employees" if self.level == "activity" else "Assigned To"
        tk.Label(f, text=emp_label, bg="#0f172a", fg="#94a3b8",
                 font=("Consolas", 9)).grid(row=row, column=0, sticky="nw", padx=8, pady=3)
        if self.level == "activity":
            current_emps = it.employees if it else []
        else:
            current_emps = it.assigned_to if it else []
        self._emp_widget = MultiEmployeeWidget(f, self.employees or [], current_emps)
        self._emp_widget.grid(row=row, column=1, sticky="ew", padx=8, pady=3)
        row += 1

        if self.level == "activity":
            # Status is computed automatically from work items
            tk.Label(f, text="Status", bg="#0f172a", fg="#94a3b8",
                     font=("Consolas", 9)).grid(row=row, column=0, sticky="w", padx=8, pady=3)
            tk.Label(f, text="Auto-calculated from work items",
                     bg="#0f172a", fg="#475569",
                     font=("Consolas", 9, "italic")).grid(row=row, column=1, sticky="w", padx=8)
            row += 1
        else:
            self._status_v = self._labeled_combo(f, "Status", row, STATUS_OPTIONS,
                                                  it.status if it else "Not Started"); row += 1

        self._priority_v = self._labeled_combo(f, "Priority", row, PRIORITY_OPTIONS,
                                                it.priority if it else "Medium"); row += 1
        self._start_v = self._labeled_date(f, "Start Date", row, it.start_date if it else ""); row += 1
        self._due_v = self._labeled_date(f, "Due Date", row, it.due_date if it else ""); row += 1

        if self.level == "workitem":
            tk.Label(f, text="Progress %", bg="#0f172a", fg="#94a3b8",
                     font=("Consolas", 9)).grid(row=row, column=0, sticky="w", padx=8, pady=3)
            self._prog_v = tk.IntVar(value=it.progress if it else 0)
            tk.Scale(f, variable=self._prog_v, from_=0, to=100, orient="horizontal",
                     bg="#0f172a", fg="#e2e8f0", troughcolor="#1e293b",
                     highlightthickness=0, length=250).grid(row=row, column=1, sticky="w", padx=8)
            row += 1

            tk.Label(f, text="Checklist", bg="#0f172a", fg="#94a3b8",
                     font=("Consolas", 9, "bold")).grid(row=row, column=0, sticky="nw", padx=8, pady=(8, 2))
            self._cl_frame = tk.Frame(f, bg="#0f172a")
            self._cl_frame.grid(row=row, column=1, sticky="ew", padx=8, pady=2)
            row += 1
            if it and it.checklist:
                for ci in it.checklist:
                    self._add_checklist_row(ci["text"], ci["done"])
            tk.Button(f, text="+ Add Checklist Item", command=self._add_checklist_row,
                      bg="#1e293b", fg="#94a3b8", relief="flat",
                      font=("Consolas", 9), cursor="hand2").grid(
                row=row, column=1, sticky="w", padx=8, pady=4)

        self._ok_cancel(self)

    def _add_checklist_row(self, text="", done=False):
        rf = tk.Frame(self._cl_frame, bg="#0f172a")
        rf.pack(fill="x", pady=1)
        done_v = tk.BooleanVar(value=done)
        text_v = tk.StringVar(value=text)
        tk.Checkbutton(rf, variable=done_v, bg="#0f172a",
                       activebackground="#0f172a").pack(side="left")
        tk.Entry(rf, textvariable=text_v, bg="#1e293b", fg="#e2e8f0",
                 insertbackground="#e2e8f0", relief="flat",
                 font=("Consolas", 9), width=28).pack(side="left", padx=4)
        tk.Button(rf, text="✕", command=rf.destroy,
                  bg="#0f172a", fg="#ef4444", relief="flat",
                  font=("Consolas", 8), cursor="hand2").pack(side="left")
        self.checklist_items.append((text_v, done_v, rf))

    def _save(self):
        t = self._title_v.get().strip()
        if not t:
            messagebox.showerror("Error", "Title is required", parent=self)
            return
        emps = self._emp_widget.get()
        if self.level == "activity":
            kwargs = dict(
                title=t,
                description=self._desc_v.get(),
                employees=emps,
                priority=self._priority_v.get(),
                start_date=self._start_v.get(),
                due_date=self._due_v.get(),
            )
            if self.item:
                for k, v in kwargs.items():
                    setattr(self.item, k, v)
                self.result = self.item
            else:
                self.result = Activity(**kwargs)
        else:
            cl = [{"text": tv.get().strip(), "done": dv.get()}
                  for tv, dv, rf in self.checklist_items
                  if rf.winfo_exists() and tv.get().strip()]
            kwargs = dict(
                title=t,
                description=self._desc_v.get(),
                assigned_to=emps,
                status=self._status_v.get(),
                priority=self._priority_v.get(),
                start_date=self._start_v.get(),
                due_date=self._due_v.get(),
                progress=self._prog_v.get(),
                checklist=cl,
            )
            if self.item:
                for k, v in kwargs.items():
                    setattr(self.item, k, v)
                self.result = self.item
            else:
                self.result = WorkItem(**kwargs)
        self.destroy()


class EmployeeDialog(BaseDialog):
    def __init__(self, parent, emp=None):
        self.emp = emp
        super().__init__(parent, "Edit Employee" if emp else "Add Employee")

    def _build(self):
        self.geometry("420x260")
        tk.Label(self, text="EMPLOYEE", bg="#0f172a", fg="#22c55e",
                 font=("Consolas", 13, "bold")).pack(pady=(16, 8))
        f = tk.Frame(self, bg="#0f172a")
        f.pack(fill="both", expand=True, padx=8)
        f.columnconfigure(1, weight=1)
        e = self.emp or {}
        self._name_v = self._labeled_entry(f, "Name", 0, e.get("name", ""))
        self._role_v = self._labeled_entry(f, "Role", 1, e.get("role", ""))
        self._email_v = self._labeled_entry(f, "Email", 2, e.get("email", ""))
        self._ok_cancel(self)

    def _save(self):
        n = self._name_v.get().strip()
        if not n:
            messagebox.showerror("Error", "Name is required", parent=self)
            return
        self.result = {"name": n, "role": self._role_v.get(), "email": self._email_v.get()}
        self.destroy()


# ─────────────────────────────────────────────
# GANTT CHART WINDOW
# ─────────────────────────────────────────────

class GanttWindow(tk.Toplevel):
    def __init__(self, parent, project: Project):
        super().__init__(parent)
        self.project = project
        self.title(f"Gantt Chart — {project.title}")
        self.configure(bg="#0f172a")
        self.geometry("1200x700")
        self._build()

    def _build(self):
        ctrl = tk.Frame(self, bg="#0f172a")
        ctrl.pack(fill="x", padx=12, pady=8)
        tk.Label(ctrl, text=f"📊 {self.project.title} — Gantt Chart",
                 bg="#0f172a", fg="#e2e8f0",
                 font=("Consolas", 13, "bold")).pack(side="left")
        tk.Button(ctrl, text="Export PNG", command=self._export_png,
                  bg="#3b82f6", fg="white", relief="flat",
                  font=("Consolas", 9), padx=10, pady=4,
                  cursor="hand2").pack(side="right", padx=4)

        self._view = tk.StringVar(value="All Levels")
        for v in ["All Levels", "Activities Only", "Work Items Only"]:
            tk.Radiobutton(ctrl, text=v, variable=self._view, value=v,
                           command=self._refresh, bg="#0f172a", fg="#94a3b8",
                           selectcolor="#1e293b", activebackground="#0f172a",
                           font=("Consolas", 9)).pack(side="left", padx=6)

        self.fig, self.ax = plt.subplots(figsize=(14, 8))
        self.fig.patch.set_facecolor("#0f172a")
        self.canvas = FigureCanvasTkAgg(self.fig, master=self)
        self.canvas.get_tk_widget().pack(fill="both", expand=True, padx=8, pady=8)
        self._refresh()

    def _refresh(self):
        self.ax.clear()
        self.ax.set_facecolor("#0a1628")
        self._draw_gantt()
        self.canvas.draw()

    def _draw_gantt(self):
        ax = self.ax
        view = self._view.get()
        p = self.project
        pc = p.color
        wi_color = lighten(pc, 0.3)

        # Build rows: each entry is a bar dict, or "gap" for inter-group spacing
        rows = []
        for g_idx, act in enumerate(p.activities):
            # Activity span: derive from work items when available so the bar
            # visually wraps all its children
            if act.work_items:
                act_start = min(parse_date(wi.start_date) for wi in act.work_items)
                act_end   = max(parse_date(wi.due_date)   for wi in act.work_items)
            else:
                act_start = parse_date(act.start_date)
                act_end   = parse_date(act.due_date)

            if view in ("All Levels", "Activities Only"):
                rows.append({
                    "label": act.title, "start": act_start, "end": act_end,
                    "color": pc, "level": 0, "prog": act.progress,
                })

            if view in ("All Levels", "Work Items Only"):
                for wi in act.work_items:
                    rows.append({
                        "label": wi.title,
                        "start": parse_date(wi.start_date),
                        "end":   parse_date(wi.due_date),
                        "color": wi_color, "level": 1, "prog": wi.progress,
                    })

            # Blank separator row between activity groups (not after the last one)
            if g_idx < len(p.activities) - 1:
                rows.append("gap")

        real = [r for r in rows if r != "gap"]
        if not real:
            ax.text(0.5, 0.5, "No tasks to display.\nAdd activities to this project.",
                    ha="center", va="center", color="#475569",
                    fontsize=12, transform=ax.transAxes)
            ax.set_xlim(0, 1)
            ax.set_ylim(-1, 1)
            return

        min_d = min(r["start"] for r in real) - timedelta(days=2)
        max_d = max(r["end"]   for r in real) + timedelta(days=4)

        # Assign y coords top-to-bottom; gap rows just consume vertical space
        n = len(rows)
        bar_h = {0: 0.58, 1: 0.34}

        for i, row in enumerate(rows):
            y = n - 1 - i
            if row == "gap":
                continue

            level = row["level"]
            start, end = row["start"], row["end"]
            prog  = row["prog"]
            color = row["color"]
            h     = bar_h[level]
            span  = max(1, (end - start).days + 1)

            # Background track
            ax.barh(y, span, left=date_to_num(start),
                    height=h, color=color, alpha=0.18, align="center")
            # Progress fill
            if prog > 0:
                ax.barh(y, max(1, int(span * prog / 100)),
                        left=date_to_num(start),
                        height=h, color=color, alpha=0.88, align="center")

            # Label
            if level == 0:
                label_text = f" ◆  {row['label'][:36]}"
                ax.text(date_to_num(min_d) + 0.3, y, label_text,
                        va="center", ha="left", color="#e2e8f0",
                        fontsize=9, fontweight="bold", fontfamily="monospace")
                # Subtle divider above each activity row (skip very first)
                if i > 0:
                    ax.axhline(y + 0.72, color="#1e3a5f",
                               linewidth=0.9, alpha=0.7, zorder=0)
            else:
                label_text = f"      · {row['label'][:34]}"
                ax.text(date_to_num(min_d) + 0.3, y, label_text,
                        va="center", ha="left", color="#94a3b8",
                        fontsize=8, fontfamily="monospace")

            # Progress % after the bar
            pct_color = "#22c55e" if prog == 100 else "#64748b"
            ax.text(date_to_num(end) + 0.4, y, f"{prog}%",
                    va="center", ha="left", color=pct_color, fontsize=7.5)

        ax.set_yticks([])
        ax.set_xlim(date_to_num(min_d), date_to_num(max_d))
        ax.set_ylim(-0.7, n - 0.3)
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %d"))
        ax.xaxis.set_major_locator(mdates.WeekdayLocator(interval=1))
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=30, ha="right",
                 color="#94a3b8", fontsize=8)
        ax.tick_params(axis="x", colors="#475569")
        ax.spines[:].set_visible(False)
        ax.set_title(f"{p.title} — Gantt Chart", color="#e2e8f0",
                     fontsize=13, pad=10, fontfamily="monospace")

        today_n = date_to_num(date.today())
        if date_to_num(min_d) < today_n < date_to_num(max_d):
            ax.axvline(today_n, color="#ef4444", linewidth=1.2,
                       linestyle="--", alpha=0.8, zorder=5)
            ax.text(today_n, n - 0.2, " Today",
                    color="#ef4444", fontsize=7.5, zorder=5)

        ax.xaxis.grid(True, color="#1e3a5f", linestyle=":", alpha=0.4)
        ax.set_axisbelow(True)
        self.fig.tight_layout(pad=1.5)

    def _export_png(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=[("PNG Image", "*.png")],
            initialfile=f"{self.project.title}_gantt.png")
        if path:
            self.fig.savefig(path, dpi=150, facecolor=self.fig.get_facecolor())
            messagebox.showinfo("Exported", f"Gantt chart saved to:\n{path}", parent=self)


# ─────────────────────────────────────────────
# EXCEL EXPORTER
# ─────────────────────────────────────────────

def _style_header(ws, row, col, text, bg="1e3a5f", fg="e2e8f0", bold=True, size=10):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Calibri", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        bottom=Side(style="thin", color="3b82f6"),
        right=Side(style="thin", color="1e3a5f"),
    )
    return cell


def _style_data(ws, row, col, value, bg="0f172a", fg="e2e8f0", bold=False, align="left"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", color=fg, bold=bold, size=9)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    return cell


def export_to_excel(project: Project, path: str):
    wb = Workbook()
    wb.remove(wb.active)

    status_colors = {
        "Not Started": "94a3b8", "In Progress": "3b82f6",
        "Completed": "22c55e", "Blocked": "ef4444", "On Hold": "f59e0b",
    }
    priority_colors = {
        "Low": "94a3b8", "Medium": "f59e0b", "High": "f97316", "Critical": "ef4444"
    }
    row_bg = ["111827", "0f172a"]

    # ── Sheet 1: Project Summary ──────────────────
    ws = wb.create_sheet("Project Summary")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGH", [28, 40, 16, 16, 24, 14, 12, 10]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = f"PROJECT PLANNER — {project.title.upper()}"
    t.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="0a1628")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    ws["A2"].value = (f"Description: {project.description}   |   "
                      f"Start: {project.start_date}   |   Due: {project.due_date}   |   "
                      f"Overall Progress: {project.progress}%")
    ws["A2"].font = Font(name="Calibri", size=9, color="94a3b8")
    ws["A2"].fill = PatternFill("solid", fgColor="0f172a")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    headers = ["Type", "Title", "Start Date", "Due Date", "Employees/Assigned", "Status", "Priority", "Progress"]
    for c, h in enumerate(headers, 1):
        _style_header(ws, 4, c, h, "1e293b", "94a3b8")
    ws.row_dimensions[4].height = 20

    data_row = 5

    def add_row(tag, title, start, end, assigned, status, priority, prog, bold=False):
        nonlocal data_row
        bg = row_bg[data_row % 2]
        _style_data(ws, data_row, 1, tag, bg, "475569", align="center")
        c = ws.cell(row=data_row, column=2, value=title)
        c.font = Font(name="Calibri", size=9, color="e2e8f0", bold=bold)
        c.fill = PatternFill("solid", fgColor=bg)
        _style_data(ws, data_row, 3, start, bg, "94a3b8", align="center")
        _style_data(ws, data_row, 4, end, bg, "94a3b8", align="center")
        _style_data(ws, data_row, 5, assigned, bg, "cbd5e1")
        sc = ws.cell(row=data_row, column=6, value=status)
        sc.font = Font(name="Calibri", size=9, color=status_colors.get(status, "e2e8f0"), bold=True)
        sc.fill = PatternFill("solid", fgColor=bg)
        pc = ws.cell(row=data_row, column=7, value=priority)
        pc.font = Font(name="Calibri", size=9, color=priority_colors.get(priority, "e2e8f0"), bold=True)
        pc.fill = PatternFill("solid", fgColor=bg)
        prog_cell = ws.cell(row=data_row, column=8, value=f"{prog}%")
        prog_cell.font = Font(name="Calibri", size=9, color="22c55e" if prog == 100 else "3b82f6")
        prog_cell.fill = PatternFill("solid", fgColor=bg)
        prog_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[data_row].height = 16
        data_row += 1

    for act in project.activities:
        add_row("ACTIVITY", act.title, act.start_date, act.due_date,
                ", ".join(act.employees), act.status, act.priority, act.progress, bold=True)
        for wi in act.work_items:
            add_row("WORK ITEM", f"  {wi.title}", wi.start_date, wi.due_date,
                    ", ".join(wi.assigned_to), wi.status, wi.priority, wi.progress)

    ws.freeze_panes = "A5"

    # ── Sheet 2: Gantt Data ──────────────────────
    wsg = wb.create_sheet("Gantt Data")
    wsg.sheet_view.showGridLines = False
    for c, h in enumerate(["Title", "Type", "Start", "End", "Duration (days)", "Progress %", "Employees"], 1):
        _style_header(wsg, 1, c, h, "1e293b", "94a3b8")
        wsg.column_dimensions[get_column_letter(c)].width = 22

    grow = 2
    for act in project.activities:
        wsg.cell(row=grow, column=1, value=act.title)
        wsg.cell(row=grow, column=2, value="Activity")
        wsg.cell(row=grow, column=3, value=act.start_date)
        wsg.cell(row=grow, column=4, value=act.due_date)
        sd, ed = parse_date(act.start_date), parse_date(act.due_date)
        wsg.cell(row=grow, column=5, value=(ed - sd).days)
        wsg.cell(row=grow, column=6, value=act.progress)
        wsg.cell(row=grow, column=7, value=", ".join(act.employees))
        for c in range(1, 8):
            wsg.cell(row=grow, column=c).fill = PatternFill("solid", fgColor="1e293b")
            wsg.cell(row=grow, column=c).font = Font(name="Calibri", color="e2e8f0", size=9, bold=True)
        grow += 1
        for wi in act.work_items:
            wsg.cell(row=grow, column=1, value=f"  {wi.title}")
            wsg.cell(row=grow, column=2, value="Work Item")
            wsg.cell(row=grow, column=3, value=wi.start_date)
            wsg.cell(row=grow, column=4, value=wi.due_date)
            sd, ed = parse_date(wi.start_date), parse_date(wi.due_date)
            wsg.cell(row=grow, column=5, value=(ed - sd).days)
            wsg.cell(row=grow, column=6, value=wi.progress)
            wsg.cell(row=grow, column=7, value=", ".join(wi.assigned_to))
            for c in range(1, 8):
                wsg.cell(row=grow, column=c).fill = PatternFill("solid", fgColor="0f172a")
                wsg.cell(row=grow, column=c).font = Font(name="Calibri", color="94a3b8", size=9)
            grow += 1

    # ── Sheet 3: Employee Workload ────────────────
    wse = wb.create_sheet("Employee Workload")
    wse.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [22, 22, 14, 14, 14]):
        wse.column_dimensions[col].width = w

    ws_title = wse.cell(row=1, column=1, value="EMPLOYEE WORKLOAD REPORT")
    ws_title.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    ws_title.fill = PatternFill("solid", fgColor="0a1628")
    wse.merge_cells("A1:E1")
    wse.row_dimensions[1].height = 28

    for c, h in enumerate(["Employee", "Assigned Tasks", "Completed", "In Progress", "Blocked"], 1):
        _style_header(wse, 2, c, h)

    emp_data: dict[str, dict] = {}
    for emp in project.employees:
        emp_data[emp["name"]] = {"tasks": 0, "completed": 0, "in_progress": 0, "blocked": 0}

    def count_emp(names, status):
        for name in names:
            if not name:
                continue
            if name not in emp_data:
                emp_data[name] = {"tasks": 0, "completed": 0, "in_progress": 0, "blocked": 0}
            emp_data[name]["tasks"] += 1
            if status == "Completed":
                emp_data[name]["completed"] += 1
            elif status == "In Progress":
                emp_data[name]["in_progress"] += 1
            elif status == "Blocked":
                emp_data[name]["blocked"] += 1

    for act in project.activities:
        count_emp(act.employees, act.status)
        for wi in act.work_items:
            count_emp(wi.assigned_to, wi.status)

    er = 3
    for name, ed_data in sorted(emp_data.items()):
        bg = row_bg[er % 2]
        _style_data(wse, er, 1, name, bg, "e2e8f0", bold=True)
        _style_data(wse, er, 2, ed_data["tasks"], bg, "3b82f6", align="center")
        _style_data(wse, er, 3, ed_data["completed"], bg, "22c55e", align="center")
        _style_data(wse, er, 4, ed_data["in_progress"], bg, "f59e0b", align="center")
        _style_data(wse, er, 5, ed_data["blocked"], bg, "ef4444", align="center")
        er += 1

    # ── Sheet 4: Checklists ───────────────────────
    wsc = wb.create_sheet("Work Item Checklists")
    wsc.sheet_view.showGridLines = False
    for c, h in enumerate(["Work Item", "Checklist Item", "Status"], 1):
        _style_header(wsc, 1, c, h)
        wsc.column_dimensions[get_column_letter(c)].width = 35

    cr = 2
    for act in project.activities:
        for wi in act.work_items:
            if wi.checklist:
                for ci in wi.checklist:
                    bg = row_bg[cr % 2]
                    _style_data(wsc, cr, 1, wi.title, bg, "e2e8f0")
                    _style_data(wsc, cr, 2, ci["text"], bg, "94a3b8")
                    done = "✓ Done" if ci["done"] else "◻ Pending"
                    _style_data(wsc, cr, 3, done, bg, "22c55e" if ci["done"] else "f59e0b")
                    cr += 1

    wb.save(path)


# ─────────────────────────────────────────────
# DASHBOARD / SNAPSHOT VIEW
# ─────────────────────────────────────────────

class SnapshotWindow(tk.Toplevel):
    def __init__(self, parent, app_data: AppData):
        super().__init__(parent)
        self.title("Project Snapshot & Dashboard")
        self.configure(bg="#0f172a")
        self.geometry("980x720")
        self.app = app_data
        self._build()

    def _build(self):
        tk.Label(self, text="📋 PROJECT SNAPSHOT & TEAM DASHBOARD",
                 bg="#0f172a", fg="#3b82f6",
                 font=("Consolas", 14, "bold")).pack(pady=(16, 4))
        tk.Label(self, text=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                 bg="#0f172a", fg="#475569",
                 font=("Consolas", 9)).pack(pady=0)

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=12, pady=12)

        t1 = tk.Frame(nb, bg="#0f172a")
        nb.add(t1, text="  Overview  ")
        self._build_overview(t1)

        t2 = tk.Frame(nb, bg="#0f172a")
        nb.add(t2, text="  Team Workload  ")
        self._build_workload(t2)

        t3 = tk.Frame(nb, bg="#0f172a")
        nb.add(t3, text="  Status Summary  ")
        self._build_status(t3)

    def _card(self, parent, title, value, sub="", color="#3b82f6"):
        f = tk.Frame(parent, bg="#1e293b", padx=16, pady=12)
        tk.Label(f, text=title, bg="#1e293b", fg="#94a3b8",
                 font=("Consolas", 8)).pack(anchor="w")
        tk.Label(f, text=str(value), bg="#1e293b", fg=color,
                 font=("Consolas", 22, "bold")).pack(anchor="w")
        if sub:
            tk.Label(f, text=sub, bg="#1e293b", fg="#475569",
                     font=("Consolas", 8)).pack(anchor="w")
        return f

    def _build_overview(self, parent):
        cf = tk.Frame(parent, bg="#0f172a")
        cf.pack(fill="x", padx=8, pady=8)

        all_activities, all_wi = [], []
        for p in self.app.projects:
            all_activities += p.activities
            for a in p.activities:
                all_wi += a.work_items

        cards = [
            ("Projects", len(self.app.projects), "", "#3b82f6"),
            ("Activities", len(all_activities), "", "#8b5cf6"),
            ("Work Items", len(all_wi), "", "#f97316"),
            ("Completed", sum(1 for w in all_wi if w.status == "Completed"), "work items", "#22c55e"),
            ("In Progress", sum(1 for w in all_wi if w.status == "In Progress"), "work items", "#3b82f6"),
            ("Blocked", sum(1 for w in all_wi if w.status == "Blocked"), "work items", "#ef4444"),
        ]
        for i, (title, val, sub, col) in enumerate(cards):
            c = self._card(cf, title, val, sub, col)
            c.grid(row=0, column=i, padx=4, pady=4, sticky="ew")
            cf.columnconfigure(i, weight=1)

        tk.Label(parent, text="PROJECT PROGRESS", bg="#0f172a", fg="#94a3b8",
                 font=("Consolas", 9, "bold")).pack(anchor="w", padx=16, pady=(12, 4))

        pf = tk.Frame(parent, bg="#0f172a")
        pf.pack(fill="both", expand=True, padx=16, pady=4)

        for p in self.app.projects:
            row_f = tk.Frame(pf, bg="#1e293b", padx=12, pady=8)
            row_f.pack(fill="x", pady=3)
            tk.Label(row_f, text=p.title, bg="#1e293b", fg="#e2e8f0",
                     font=("Consolas", 10, "bold"), width=30, anchor="w").pack(side="left")
            bar_f = tk.Frame(row_f, bg="#0f172a", height=14, width=300)
            bar_f.pack(side="left", padx=8)
            bar_f.pack_propagate(False)
            prog = p.progress
            tk.Frame(bar_f, bg=p.color, width=max(1, int(300 * prog / 100)), height=14).place(x=0, y=0)
            tk.Label(row_f, text=f"{prog}%", bg="#1e293b",
                     fg="#94a3b8", font=("Consolas", 9)).pack(side="left", padx=4)
            completed = sum(1 for a in p.activities if a.status == "Completed")
            tk.Label(row_f,
                     text=f"{completed}/{len(p.activities)} activities done",
                     bg="#1e293b", fg="#475569",
                     font=("Consolas", 8)).pack(side="right")

    def _build_workload(self, parent):
        emp_data: dict[str, dict] = {}

        for p in self.app.projects:
            def count(names, status, pname):
                for name in names:
                    if not name:
                        continue
                    if name not in emp_data:
                        emp_data[name] = {
                            "tasks": 0, "completed": 0, "in_progress": 0,
                            "blocked": 0, "projects": set()
                        }
                    emp_data[name]["tasks"] += 1
                    emp_data[name]["projects"].add(pname)
                    if status == "Completed":
                        emp_data[name]["completed"] += 1
                    elif status == "In Progress":
                        emp_data[name]["in_progress"] += 1
                    elif status == "Blocked":
                        emp_data[name]["blocked"] += 1

            for act in p.activities:
                count(act.employees, act.status, p.title)
                for wi in act.work_items:
                    count(wi.assigned_to, wi.status, p.title)

        if not emp_data:
            tk.Label(parent, text="No employees assigned to tasks yet.",
                     bg="#0f172a", fg="#475569",
                     font=("Consolas", 11)).pack(expand=True)
            return

        cols = ("Employee", "Total Tasks", "Completed", "In Progress", "Blocked", "Projects")
        tree = ttk.Treeview(parent, columns=cols, show="headings", height=20)
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", background="#0f172a", fieldbackground="#0f172a",
                        foreground="#e2e8f0", rowheight=28, font=("Consolas", 9))
        style.configure("Treeview.Heading", background="#1e293b", foreground="#94a3b8",
                        font=("Consolas", 9, "bold"))

        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=130, anchor="center")
        tree.column("Employee", width=180, anchor="w")
        tree.column("Projects", width=220, anchor="w")

        for name, d in sorted(emp_data.items()):
            tree.insert("", "end", values=(
                name, d["tasks"], d["completed"], d["in_progress"],
                d["blocked"], ", ".join(sorted(d["projects"]))
            ))
        tree.pack(fill="both", expand=True, padx=12, pady=12)

    def _build_status(self, parent):
        act_counts = {s: 0 for s in STATUS_OPTIONS}
        wi_counts = {s: 0 for s in STATUS_OPTIONS}

        for p in self.app.projects:
            for act in p.activities:
                act_counts[act.status] = act_counts.get(act.status, 0) + 1
                for wi in act.work_items:
                    wi_counts[wi.status] = wi_counts.get(wi.status, 0) + 1

        fig, axes = plt.subplots(1, 2, figsize=(10, 4))
        fig.patch.set_facecolor("#0f172a")

        for ax, counts, title in [(axes[0], act_counts, "Activities"), (axes[1], wi_counts, "Work Items")]:
            ax.set_facecolor("#0a1628")
            vals = [v for v in counts.values() if v > 0]
            labels = [k for k, v in counts.items() if v > 0]
            colors = [STATUS_COLORS.get(l, "#94a3b8") for l in labels]
            if vals:
                ax.pie(vals, labels=labels, colors=colors, autopct="%1.0f%%",
                       textprops={"color": "#e2e8f0", "fontsize": 8}, pctdistance=0.7)
            ax.set_title(title, color="#94a3b8", fontsize=10, fontfamily="monospace")

        fig.suptitle("Status Distribution by Level", color="#e2e8f0",
                     fontsize=12, fontfamily="monospace")
        canvas = FigureCanvasTkAgg(fig, master=parent)
        canvas.get_tk_widget().pack(fill="both", expand=True, padx=8, pady=8)
        canvas.draw()


# ─────────────────────────────────────────────
# MAIN APPLICATION
# ─────────────────────────────────────────────

class ProjectPlannerApp(tk.Tk):
    SAVE_FILE = "project_planner_data.json"

    def __init__(self):
        super().__init__()
        self.title("Project Planner")
        self.geometry("1400x820")
        self.configure(bg="#0a1628")
        self.data = AppData()
        self._selected_project: Optional[Project] = None
        self._selected_activity: Optional[Activity] = None
        self._setup_style()
        self._build_ui()
        self._load()
        self._refresh_projects()

    def _setup_style(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TNotebook", background="#0a1628", borderwidth=0)
        style.configure("TNotebook.Tab", background="#0f172a", foreground="#94a3b8",
                        font=("Consolas", 9), padding=[12, 5])
        style.map("TNotebook.Tab",
                  background=[("selected", "#1e293b")],
                  foreground=[("selected", "#e2e8f0")])
        style.configure("Treeview", background="#0f172a", fieldbackground="#0f172a",
                        foreground="#e2e8f0", rowheight=26, font=("Consolas", 9))
        style.configure("Treeview.Heading", background="#1e293b", foreground="#94a3b8",
                        font=("Consolas", 9, "bold"))
        style.map("Treeview", background=[("selected", "#1e3a5f")])
        style.configure("TScrollbar", background="#1e293b", troughcolor="#0f172a",
                        arrowcolor="#475569", borderwidth=0)

    def _btn(self, parent, text, cmd, color="#1e293b", fg="#94a3b8", small=False):
        sz = 8 if small else 9
        return tk.Button(parent, text=text, command=cmd,
                         bg=color, fg=fg, relief="flat",
                         font=("Consolas", sz), padx=8, pady=4,
                         cursor="hand2", activebackground=color,
                         activeforeground="white")

    def _build_ui(self):
        top = tk.Frame(self, bg="#020817", height=48)
        top.pack(fill="x", side="top")
        top.pack_propagate(False)

        tk.Label(top, text="◆ PROJECT PLANNER",
                 bg="#020817", fg="#3b82f6",
                 font=("Consolas", 14, "bold")).pack(side="left", padx=20, pady=10)

        for text, cmd, color in [
            ("＋ New Project", self._new_project, "#3b82f6"),
            ("📊 Dashboard", self._open_snapshot, "#8b5cf6"),
            ("👥 Employees", self._manage_employees, "#14b8a6"),
            ("💾 Save", self._save, "#334155"),
            ("📂 Load", self._load_from, "#334155"),
        ]:
            self._btn(top, text, cmd, color,
                      fg="white" if color != "#334155" else "#94a3b8").pack(
                side="left", padx=3, pady=10)

        main = tk.Frame(self, bg="#0a1628")
        main.pack(fill="both", expand=True)

        # LEFT: Project list
        sidebar = tk.Frame(main, bg="#0f172a", width=220)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        tk.Label(sidebar, text="PROJECTS", bg="#0f172a", fg="#475569",
                 font=("Consolas", 8, "bold")).pack(anchor="w", padx=12, pady=(12, 4))

        self._proj_listbox = tk.Listbox(
            sidebar, bg="#0f172a", fg="#e2e8f0", selectbackground="#1e3a5f",
            selectforeground="#e2e8f0", relief="flat", font=("Consolas", 10),
            borderwidth=0, highlightthickness=0, activestyle="none"
        )
        self._proj_listbox.pack(fill="both", expand=True, padx=4)
        self._proj_listbox.bind("<<ListboxSelect>>", self._on_project_select)

        pb = tk.Frame(sidebar, bg="#0f172a")
        pb.pack(fill="x", padx=6, pady=6)
        self._btn(pb, "Edit", self._edit_project, small=True).pack(side="left", padx=2)
        self._btn(pb, "Delete", self._delete_project, "#334155", "#ef4444", small=True).pack(side="left", padx=2)

        # RIGHT: tabbed content
        right = tk.Frame(main, bg="#0a1628")
        right.pack(side="left", fill="both", expand=True)

        self._nb = ttk.Notebook(right)
        self._nb.pack(fill="both", expand=True, padx=4, pady=4)

        self._tab_tasks = tk.Frame(self._nb, bg="#0a1628")
        self._nb.add(self._tab_tasks, text="  Tasks  ")
        self._build_task_tab()

        self._tab_gantt = tk.Frame(self._nb, bg="#0a1628")
        self._nb.add(self._tab_gantt, text="  Gantt Chart  ")
        self._build_gantt_tab()

        self._tab_emp = tk.Frame(self._nb, bg="#0a1628")
        self._nb.add(self._tab_emp, text="  Team  ")
        self._build_emp_tab()

        self._nb.bind("<<NotebookTabChanged>>", self._on_tab_change)

    def _build_task_tab(self):
        parent = self._tab_tasks
        pane = tk.PanedWindow(parent, orient="horizontal",
                              bg="#0a1628", sashwidth=4, sashrelief="flat")
        pane.pack(fill="both", expand=True)

        p1 = tk.Frame(pane, bg="#0a1628")
        pane.add(p1, width=480)
        self._build_activity_pane(p1)

        p2 = tk.Frame(pane, bg="#0a1628")
        pane.add(p2, width=480)
        self._build_workitem_pane(p2)

    def _pane_header(self, parent, title, color, add_cmd):
        h = tk.Frame(parent, bg="#0f172a")
        h.pack(fill="x")
        tk.Label(h, text=title, bg="#0f172a", fg=color,
                 font=("Consolas", 10, "bold")).pack(side="left", padx=12, pady=8)
        self._btn(h, "+ Add", add_cmd, color, "white", small=True).pack(
            side="right", padx=8, pady=6)

    def _build_activity_pane(self, parent):
        self._pane_header(parent, "◆ ACTIVITIES", "#8b5cf6", self._add_activity)
        cols = ("title", "employees", "status", "progress", "priority")
        self._act_tree = ttk.Treeview(parent, columns=cols, show="headings", selectmode="browse")
        self._act_tree.heading("title", text="Title")
        self._act_tree.heading("employees", text="Employees")
        self._act_tree.heading("status", text="Status")
        self._act_tree.heading("progress", text="%")
        self._act_tree.heading("priority", text="Pri")
        self._act_tree.column("title", width=140)
        self._act_tree.column("employees", width=130)
        self._act_tree.column("status", width=90)
        self._act_tree.column("progress", width=40, anchor="center")
        self._act_tree.column("priority", width=50, anchor="center")
        sb = ttk.Scrollbar(parent, orient="vertical", command=self._act_tree.yview)
        self._act_tree.configure(yscrollcommand=sb.set)
        self._act_tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self._act_tree.bind("<<TreeviewSelect>>", self._on_activity_select)
        self._act_tree.bind("<Double-1>", lambda e: self._edit_activity())

        bf = tk.Frame(parent, bg="#0a1628")
        bf.pack(fill="x", padx=6, pady=4)
        self._btn(bf, "Edit", self._edit_activity, small=True).pack(side="left", padx=2)
        self._btn(bf, "Delete", self._delete_activity, "#334155", "#ef4444", small=True).pack(side="left")

    def _build_workitem_pane(self, parent):
        self._pane_header(parent, "· WORK ITEMS", "#f97316", self._add_workitem)
        cols = ("title", "assigned", "status", "progress", "priority")
        self._wi_tree = ttk.Treeview(parent, columns=cols, show="headings", selectmode="browse")
        self._wi_tree.heading("title", text="Title")
        self._wi_tree.heading("assigned", text="Assigned To")
        self._wi_tree.heading("status", text="Status")
        self._wi_tree.heading("progress", text="%")
        self._wi_tree.heading("priority", text="Pri")
        self._wi_tree.column("title", width=140)
        self._wi_tree.column("assigned", width=130)
        self._wi_tree.column("status", width=90)
        self._wi_tree.column("progress", width=40, anchor="center")
        self._wi_tree.column("priority", width=50, anchor="center")
        sb = ttk.Scrollbar(parent, orient="vertical", command=self._wi_tree.yview)
        self._wi_tree.configure(yscrollcommand=sb.set)
        self._wi_tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self._wi_tree.bind("<Double-1>", lambda e: self._edit_workitem())

        bf = tk.Frame(parent, bg="#0a1628")
        bf.pack(fill="x", padx=6, pady=4)
        self._btn(bf, "Edit", self._edit_workitem, small=True).pack(side="left", padx=2)
        self._btn(bf, "Delete", self._delete_workitem, "#334155", "#ef4444", small=True).pack(side="left")
        self._btn(bf, "Checklist", self._view_checklist, "#334155", "#94a3b8", small=True).pack(side="left", padx=2)

    def _build_gantt_tab(self):
        parent = self._tab_gantt
        ctrl = tk.Frame(parent, bg="#0a1628")
        ctrl.pack(fill="x", padx=12, pady=8)
        tk.Label(ctrl, text="Select a project then open Gantt chart:",
                 bg="#0a1628", fg="#94a3b8",
                 font=("Consolas", 9)).pack(side="left")
        self._btn(ctrl, "Open Gantt Chart ↗", self._open_gantt,
                  "#3b82f6", "white").pack(side="left", padx=8)
        self._btn(ctrl, "Export to Excel", self._export_excel,
                  "#22c55e", "white").pack(side="left", padx=4)

        self._gantt_fig, self._gantt_ax = plt.subplots(figsize=(12, 6))
        self._gantt_fig.patch.set_facecolor("#0f172a")
        self._gantt_canvas = FigureCanvasTkAgg(self._gantt_fig, master=parent)
        self._gantt_canvas.get_tk_widget().pack(fill="both", expand=True, padx=8, pady=4)

    def _build_emp_tab(self):
        parent = self._tab_emp
        ctrl = tk.Frame(parent, bg="#0a1628")
        ctrl.pack(fill="x", padx=12, pady=8)
        tk.Label(ctrl, text="TEAM MEMBERS", bg="#0a1628", fg="#14b8a6",
                 font=("Consolas", 11, "bold")).pack(side="left")
        self._btn(ctrl, "+ Add Employee", self._add_project_employee,
                  "#14b8a6", "white").pack(side="left", padx=8)
        self._btn(ctrl, "Remove", self._remove_project_employee,
                  "#334155", "#ef4444").pack(side="left")

        cols = ("name", "role", "email")
        self._emp_tree = ttk.Treeview(parent, columns=cols, show="headings", height=20)
        self._emp_tree.heading("name", text="Name")
        self._emp_tree.heading("role", text="Role")
        self._emp_tree.heading("email", text="Email")
        self._emp_tree.column("name", width=200)
        self._emp_tree.column("role", width=180)
        self._emp_tree.column("email", width=220)
        sb = ttk.Scrollbar(parent, orient="vertical", command=self._emp_tree.yview)
        self._emp_tree.configure(yscrollcommand=sb.set)
        self._emp_tree.pack(side="left", fill="both", expand=True, padx=(12, 0), pady=8)
        sb.pack(side="right", fill="y", pady=8, padx=(0, 12))

    # ── Refresh ──────────────────────────────────

    def _sync_activity_row(self):
        """Update the activity row in the treeview without resetting selection."""
        act = self._selected_activity
        if not act:
            return
        try:
            self._act_tree.item(act.id, values=(
                act.title,
                ", ".join(act.employees) or "—",
                act.status,
                f"{act.progress}%",
                act.priority,
            ))
        except tk.TclError:
            pass

    def _refresh_projects(self):
        self._proj_listbox.delete(0, "end")
        for p in self.data.projects:
            self._proj_listbox.insert("end", f"  {p.title}  ({p.progress}%)")

    def _refresh_activities(self):
        for item in self._act_tree.get_children():
            self._act_tree.delete(item)
        if not self._selected_project:
            self._refresh_workitems()
            return
        for act in self._selected_project.activities:
            self._act_tree.insert("", "end", iid=act.id,
                                   values=(act.title,
                                           ", ".join(act.employees) or "—",
                                           act.status,
                                           f"{act.progress}%",
                                           act.priority))
        self._refresh_workitems()

    def _refresh_workitems(self):
        for item in self._wi_tree.get_children():
            self._wi_tree.delete(item)
        if not self._selected_activity:
            return
        for wi in self._selected_activity.work_items:
            self._wi_tree.insert("", "end", iid=wi.id,
                                  values=(wi.title,
                                          ", ".join(wi.assigned_to) or "—",
                                          wi.status,
                                          f"{wi.progress}%",
                                          wi.priority))

    def _refresh_employees(self):
        for item in self._emp_tree.get_children():
            self._emp_tree.delete(item)
        if not self._selected_project:
            return
        for e in self._selected_project.employees:
            self._emp_tree.insert("", "end", values=(e["name"], e["role"], e["email"]))

    def _refresh_inline_gantt(self):
        self._gantt_ax.clear()
        self._gantt_ax.set_facecolor("#0a1628")
        p = self._selected_project
        if not p:
            self._gantt_ax.text(0.5, 0.5, "Select a project to preview Gantt",
                                ha="center", va="center", color="#475569",
                                fontsize=11, transform=self._gantt_ax.transAxes)
            self._gantt_canvas.draw()
            return
        gw = GanttWindow.__new__(GanttWindow)
        gw.project = p
        gw.ax = self._gantt_ax
        gw.fig = self._gantt_fig
        gw._view = tk.StringVar(value="All Levels")
        gw._draw_gantt()
        self._gantt_canvas.draw()

    # ── Selection handlers ───────────────────────

    def _on_project_select(self, event=None):
        sel = self._proj_listbox.curselection()
        if not sel:
            return
        idx = sel[0]
        if idx < len(self.data.projects):
            self._selected_project = self.data.projects[idx]
            self._selected_activity = None
            self._refresh_activities()
            self._refresh_employees()
            self._refresh_inline_gantt()

    def _on_activity_select(self, event=None):
        sel = self._act_tree.selection()
        if not sel or not self._selected_project:
            return
        iid = sel[0]
        self._selected_activity = next(
            (a for a in self._selected_project.activities if a.id == iid), None)
        self._refresh_workitems()

    def _on_tab_change(self, event=None):
        if self._nb.index(self._nb.select()) == 1:
            self._refresh_inline_gantt()

    # ── Project CRUD ─────────────────────────────

    def _new_project(self):
        dlg = ProjectDialog(self)
        if dlg.result:
            self.data.projects.append(dlg.result)
            self._refresh_projects()
            self._save()

    def _edit_project(self):
        if not self._selected_project:
            return
        dlg = ProjectDialog(self, self._selected_project)
        if dlg.result:
            self._refresh_projects()
            self._save()

    def _delete_project(self):
        if not self._selected_project:
            return
        if messagebox.askyesno("Delete", f"Delete project '{self._selected_project.title}'?"):
            self.data.projects.remove(self._selected_project)
            self._selected_project = None
            self._selected_activity = None
            self._refresh_projects()
            self._refresh_activities()
            self._save()

    # ── Activity CRUD ─────────────────────────────

    def _employees_list(self):
        if self._selected_project:
            return [e["name"] for e in self._selected_project.employees]
        return self.data.all_employees()

    def _add_activity(self):
        if not self._selected_project:
            messagebox.showwarning("Warning", "Select a project first")
            return
        dlg = TaskDialog(self, "activity", employees=self._employees_list())
        if dlg.result:
            self._selected_project.activities.append(dlg.result)
            self._refresh_activities()
            self._save()

    def _edit_activity(self):
        sel = self._act_tree.selection()
        if not sel or not self._selected_project:
            return
        iid = sel[0]
        act = next((a for a in self._selected_project.activities if a.id == iid), None)
        if act:
            dlg = TaskDialog(self, "activity", act, employees=self._employees_list())
            if dlg.result:
                self._refresh_activities()
                self._save()

    def _delete_activity(self):
        sel = self._act_tree.selection()
        if not sel or not self._selected_project:
            return
        iid = sel[0]
        act = next((a for a in self._selected_project.activities if a.id == iid), None)
        if act and messagebox.askyesno("Delete", f"Delete activity '{act.title}'?"):
            self._selected_project.activities.remove(act)
            if self._selected_activity == act:
                self._selected_activity = None
            self._refresh_activities()
            self._save()

    # ── Work Item CRUD ────────────────────────────

    def _add_workitem(self):
        if not self._selected_activity:
            messagebox.showwarning("Warning", "Select an activity first")
            return
        dlg = TaskDialog(self, "workitem", employees=self._employees_list())
        if dlg.result:
            self._selected_activity.work_items.append(dlg.result)
            self._refresh_workitems()
            self._sync_activity_row()
            self._save()

    def _edit_workitem(self):
        sel = self._wi_tree.selection()
        if not sel or not self._selected_activity:
            return
        iid = sel[0]
        wi = next((w for w in self._selected_activity.work_items if w.id == iid), None)
        if wi:
            dlg = TaskDialog(self, "workitem", wi, employees=self._employees_list())
            if dlg.result:
                self._refresh_workitems()
                self._sync_activity_row()
                self._save()

    def _delete_workitem(self):
        sel = self._wi_tree.selection()
        if not sel or not self._selected_activity:
            return
        iid = sel[0]
        wi = next((w for w in self._selected_activity.work_items if w.id == iid), None)
        if wi and messagebox.askyesno("Delete", f"Delete work item '{wi.title}'?"):
            self._selected_activity.work_items.remove(wi)
            self._refresh_workitems()
            self._sync_activity_row()
            self._save()

    def _view_checklist(self):
        sel = self._wi_tree.selection()
        if not sel or not self._selected_activity:
            return
        iid = sel[0]
        wi = next((w for w in self._selected_activity.work_items if w.id == iid), None)
        if not wi:
            return
        win = tk.Toplevel(self)
        win.title(f"Checklist — {wi.title}")
        win.configure(bg="#0f172a")
        win.geometry("380x440")
        win.grab_set()
        tk.Label(win, text=f"✅ Checklist: {wi.title}",
                 bg="#0f172a", fg="#f97316",
                 font=("Consolas", 11, "bold")).pack(pady=12)
        if not wi.checklist:
            tk.Label(win, text="No checklist items.",
                     bg="#0f172a", fg="#475569",
                     font=("Consolas", 10)).pack(expand=True)
        else:
            for ci in wi.checklist:
                rf = tk.Frame(win, bg="#1e293b", padx=10, pady=6)
                rf.pack(fill="x", padx=12, pady=2)
                mark = "✓" if ci["done"] else "◻"
                col = "#22c55e" if ci["done"] else "#94a3b8"
                tk.Label(rf, text=f"{mark}  {ci['text']}",
                         bg="#1e293b", fg=col,
                         font=("Consolas", 9)).pack(anchor="w")
        done = sum(1 for ci in wi.checklist if ci["done"])
        tk.Label(win, text=f"{done}/{len(wi.checklist)} items complete",
                 bg="#0f172a", fg="#475569",
                 font=("Consolas", 9)).pack(pady=8)
        tk.Button(win, text="Close", command=win.destroy,
                  bg="#334155", fg="#94a3b8", relief="flat",
                  font=("Consolas", 9), padx=16, pady=6).pack(pady=4)

    # ── Employee management ────────────────────────

    def _add_project_employee(self):
        if not self._selected_project:
            messagebox.showwarning("Warning", "Select a project first")
            return
        dlg = EmployeeDialog(self)
        if dlg.result:
            self._selected_project.employees.append(dlg.result)
            self._refresh_employees()
            self._save()

    def _remove_project_employee(self):
        if not self._selected_project:
            return
        sel = self._emp_tree.selection()
        if not sel:
            return
        idx = self._emp_tree.index(sel[0])
        emp = self._selected_project.employees[idx]
        if messagebox.askyesno("Remove", f"Remove {emp['name']} from project?"):
            self._selected_project.employees.pop(idx)
            self._refresh_employees()
            self._save()

    def _manage_employees(self):
        win = tk.Toplevel(self)
        win.title("Global Employee Pool")
        win.configure(bg="#0f172a")
        win.geometry("500x420")
        win.grab_set()
        tk.Label(win, text="👥 GLOBAL EMPLOYEES",
                 bg="#0f172a", fg="#14b8a6",
                 font=("Consolas", 12, "bold")).pack(pady=12)
        cols = ("name", "role", "email")
        tree = ttk.Treeview(win, columns=cols, show="headings", height=14)
        tree.heading("name", text="Name")
        tree.heading("role", text="Role")
        tree.heading("email", text="Email")
        tree.pack(fill="both", expand=True, padx=12, pady=4)

        def refresh():
            for i in tree.get_children():
                tree.delete(i)
            for e in self.data.global_employees:
                tree.insert("", "end", values=(e["name"], e["role"], e["email"]))

        refresh()
        bf = tk.Frame(win, bg="#0f172a")
        bf.pack(fill="x", padx=12, pady=6)

        def add():
            dlg = EmployeeDialog(win)
            if dlg.result:
                self.data.global_employees.append(dlg.result)
                refresh()
                self._save()

        def remove():
            sel = tree.selection()
            if sel:
                self.data.global_employees.pop(tree.index(sel[0]))
                refresh()
                self._save()

        self._btn(bf, "+ Add", add, "#14b8a6", "white").pack(side="left", padx=4)
        self._btn(bf, "Remove", remove, "#334155", "#ef4444").pack(side="left")

    # ── Gantt & Export ─────────────────────────────

    def _open_gantt(self):
        if not self._selected_project:
            messagebox.showwarning("Warning", "Select a project first")
            return
        GanttWindow(self, self._selected_project)

    def _export_excel(self):
        if not self._selected_project:
            messagebox.showwarning("Warning", "Select a project first")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile=f"{self._selected_project.title}_plan.xlsx")
        if path:
            try:
                export_to_excel(self._selected_project, path)
                messagebox.showinfo("Exported", f"Project plan exported to:\n{path}")
            except Exception as e:
                messagebox.showerror("Error", f"Export failed:\n{e}")

    def _open_snapshot(self):
        SnapshotWindow(self, self.data)

    # ── Save / Load ────────────────────────────────

    def _save(self):
        self.data.save(self.SAVE_FILE)

    def _load(self):
        if os.path.exists(self.SAVE_FILE):
            try:
                self.data.load(self.SAVE_FILE)
            except Exception as e:
                print(f"Load error: {e}")

    def _load_from(self):
        path = filedialog.askopenfilename(
            filetypes=[("JSON", "*.json"), ("All Files", "*.*")])
        if path:
            try:
                self.data.load(path)
                self.SAVE_FILE = path
                self._selected_project = None
                self._selected_activity = None
                self._refresh_projects()
                self._refresh_activities()
            except Exception as e:
                messagebox.showerror("Error", f"Load failed:\n{e}")


# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────

if __name__ == "__main__":
    app = ProjectPlannerApp()
    app.mainloop()
